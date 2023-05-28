import os
import json
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
from tqdm import tqdm

def get_col_rows(excel_file) -> list:
    """Get the values of each table/spreadsheet in a column and row-wise format.

    Args:
        excel_file (str): This is the excel file that contains the table

    Returns:
        list: A list of list containing the values in column-wise and row-wise format.
    """
    table = load_workbook(filename=f"raw-data/{excel_file}")["Daily"]

    table_total_rows = table.max_row

    table_col_range = table.iter_cols(min_col=2, max_col=27, min_row=1, max_row=table_total_rows, values_only=True)
    table_row_range = table.iter_rows(min_row=1, max_row=table_total_rows, min_col=1, max_col=27, values_only=True)

    table_cols = [col for col in table_col_range]
    table_rows = [row for row in table_row_range]
    
    return table_cols, table_rows

def extract_ridership(col_list) -> dict:
    """ Extracts the ridership values from a list of list and puts it in a dictionary.
        This function already partially cleans the data by removing values that are not integers (i.e, str and None types)

    Args:
        col_list (list): This is a list of list that contains ridership values for each station—grouped by columns.

    Returns:
        dict: This a dictionary containing the ridership numbers as its values and the station names as its keys
    """
    
    station_riders_dict = {}
    
    for idx_col, col in enumerate(col_list):
        
        value_list = []
        str_list = []
        
        for idx_cell, cell in enumerate(col):
            reference_cell = col_list[idx_col-1][idx_cell]
            
            if cell is None:           
                 if isinstance(reference_cell, int):
                    value_list.append(np.nan)
                
            if isinstance(cell, int):
                value_list.append(cell)
            
            if isinstance(cell, str):
                str_list.append(cell)   
        
        station_name = str_list[0]
        
        station_riders_dict[station_name] = value_list
        
    del station_riders_dict["Exit"]
    
    return station_riders_dict

def extract_time(row_list) -> list:
    """ Extracts the hour values from a list of list and puts it in a list.
        This function only extracts the hour values that have ridership values associated with it.

    Args:
        row_list (list): This is a list of list that contains values for all stations—grouped by rows.

    Returns:
        list: This a list containing the hours that have values associated with it (i.e., non-empty entries)
    """
    
    time_list = []
    
    for row in row_list:
        if isinstance(row[5], int):
            time_list.append(row[0])
    
    return time_list

def generate_date(row_list, excel_file) -> list:
    """Generates date associated with the entries from the table.
       The assumption of this function is that if the cell contains the string, "Entry",
       then it means that following set of values belong to a new day/date.
        
    Args:
        row_list (list): This is a list of list that contains values for all stations—grouped by rows.

    Returns:
        list: This is a list containing the dates that have values associated with it (i.e., non-empty entries)
    """
    
    dates_list = []
    day = 0
    year_month = excel_file.split(".")[0]
    
    for row in row_list:
        if row[1] == "Entry":
            day += 1
        
        if isinstance(row[5], int):
            dates_list.append(f"{year_month}-{day}")
        
    return dates_list

def troubleshoot(time_list, dates_list, station_riders_dict,excel_file):
    """Provides a summary of the dataset to spot discrepancies.

    Args:
        time_list (list): This is a list containing all the hour elements extracted from the dataset
        dates_list (list): This ia list containing all the date elements generated from the dataset 
        station_riders_dict (_type_): This is a dictionary of a list containing the ridership values of each station.
        excel_file (_type_): This is the file name of the excel workbook
    """
    
    hours_count = len(time_list)
    dates_count = len(dates_list)
    station_count = {}
    date_written = datetime.strftime(datetime.now(), "%d%B%Y")
    
    for station, values in station_riders_dict.items():
        station_count[f"{station}"] = hours_count - len(values)
    
    with open(f"log/{date_written}-log.txt", "a") as log:
        log.write(f"{excel_file} Summary\n")
        log.write(f"Time element count: {hours_count}\n")   
        log.write(f"Date element count: {dates_count}\n") 
        log.write(f"Station element error count: {station_count}\n\n")
        
def compile_values(file_list):
    """Compiles the values from all the tables in the collection of spreadsheets.

    Args:
        file_list (sts): This is the file name/path of the data/table/spreadsheet.

    Returns:
        list and dict: Returns the extracted values from all the columns in the table.
    """
    
    time_list = []
    dates_list = []
    station_dict = {}

    for excel_file in tqdm(file_list):
        try:
            table_cols, table_rows = get_col_rows(excel_file)
            
            times = extract_time(table_rows)
            dates = generate_date(table_rows, excel_file)
            station_riders_dict = extract_ridership(table_cols)
            
            for time in times:
                time_list.append(time)
            
            for date in dates:
                dates_list.append(date)
            
            if station_dict: 
                for station, values in station_riders_dict.items():
                    station_dict[f"{station}"] += values
            else:
                for station, values in station_riders_dict.items():
                    station_dict[f"{station}"] = values
            
            troubleshoot(times, dates, station_riders_dict, excel_file)
            
        except ValueError:
            print(f"{excel_file} skipped.")
            pass
        
        
        
    return time_list, dates_list, station_dict

def merge_to_json(time_list, dates_list, station_dict):
    """This merges all the column and its values into a dictionary and saves it as a JSON file.

    Args:
        time_list (list): Contains all the time elements
        dates_list (list): Contains all the date elements
        station_dict (dict): Contains all the rider count for each station

    """
    
    metro_dict = {}
    
    metro_dict["date"] = dates_list
    metro_dict["time"] = time_list
    
    for station, values in station_dict.items():
        metro_dict[f"{station}"] = values
    
    metro_object = json.dumps(metro_dict, indent=4)
    
    with open("cleaned-data/mrt_data.json", "w") as mrt_data:
        mrt_data.write(metro_object)
        
def main():
    excel_file_list = os.listdir("raw-data")
    
    time, dates, stations = compile_values(excel_file_list)
    
    json_metro_data = merge_to_json(time, dates, stations)
    
    print(f"Summary: time ({len(time)}), dates ({len(dates)}), stations ({[(len(dates) - len(values)) for station, values in stations.items()]})")
    
#  main directive running the program
if __name__ == "__main__":
    main()