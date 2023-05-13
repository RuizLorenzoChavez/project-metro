from openpyxl import load_workbook
import os

#  list of all available exceel workbooks
excel_file_list = os.listdir("data")


daily = load_workbook(filename=f"data/{excel_file_list[0]}")["Daily"]

daily_total_row = daily.max_row

daily_table_range = daily.iter_rows(min_row=1, max_row=daily_total_row,min_col=1, max_col=27)

daily_rows = [row for row in daily_table_range]

